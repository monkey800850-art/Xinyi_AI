import io
import json
from datetime import datetime
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.db import get_engine
from app.services.ar_ap_service import get_aging_report
from app.services.aux_reports_service import get_aux_balance, get_aux_ledger
from app.services.bank_reconcile_service import get_reconcile_report
from app.services.payment_service import list_payments
from app.services.tax_service import get_tax_summary
from app.services.trial_balance_service import get_trial_balance
from app.services.ledger_service import get_subject_ledger
from app.services.asset_reports_service import (
    get_asset_ledger,
    get_depreciation_detail,
    get_depreciation_summary,
)
from app.services.audit_service import log_audit


class ExportError(RuntimeError):
    def __init__(self, message: str):
        super().__init__(message)


def _format_amount(value) -> str:
    try:
        return f"{float(value):.2f}"
    except Exception:
        return str(value)


def _write_sheet(ws, title: str, filters: Dict[str, object], headers: List[str], rows: List[List[object]]):
    ws.append([title])
    ws.append([f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    filter_parts = []
    for k, v in filters.items():
        if v in (None, ""):
            continue
        filter_parts.append(f"{k}={v}")
    ws.append(["筛选条件: " + ("; ".join(filter_parts) if filter_parts else "无")])
    ws.append([])
    ws.append(headers)

    for row in rows:
        ws.append(row)

    # basic column width
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18


def _audit_export(report_key: str, book_id, filters: Dict[str, object], file_name: str, operator: str):
    engine = get_engine()
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO export_audit_logs (report_key, book_id, filters, file_name, operator)
                VALUES (:report_key, :book_id, :filters, :file_name, :operator)
                """
            ),
            {
                "report_key": report_key,
                "book_id": book_id,
                "filters": json.dumps(filters, ensure_ascii=False),
                "file_name": file_name,
                "operator": operator or "",
            },
        )
    log_audit(
        "export",
        "export",
        "report",
        None,
        operator,
        "",
        {"report_key": report_key, "file_name": file_name, "filters": filters},
    )


def _build_workbook(report_key: str, params: Dict[str, str]) -> Tuple[Workbook, Dict[str, object]]:
    wb = Workbook()
    ws = wb.active

    if report_key == "trial_balance":
        data = get_trial_balance(params)
        filters = {
            "book_id": data["book_id"],
            "start_date": data["start_date"],
            "end_date": data["end_date"],
        }
        headers = ["科目编码", "科目名称", "期初余额", "本期借方发生额", "本期贷方发生额", "期末余额"]
        rows = [
            [
                item["code"],
                item["name"],
                _format_amount(item["opening_balance"]),
                _format_amount(item["period_debit"]),
                _format_amount(item["period_credit"]),
                _format_amount(item["ending_balance"]),
            ]
            for item in data.get("items", [])
        ]
        _write_sheet(ws, "发生余额表", filters, headers, rows)
        return wb, filters

    if report_key == "subject_ledger":
        data = get_subject_ledger(params)
        filters = {
            "book_id": data["book_id"],
            "subject_code": data["subject_code"],
            "subject_name": data["subject_name"],
            "start_date": data["start_date"],
            "end_date": data["end_date"],
            "summary": params.get("summary", ""),
            "direction": params.get("direction", ""),
        }
        headers = ["日期", "凭证号", "行号", "摘要", "借方", "贷方", "备注"]
        rows = [
            [
                item["voucher_date"],
                f"{item['voucher_word']}{item['voucher_no']}",
                item["line_no"],
                item["summary"],
                _format_amount(item["debit"]),
                _format_amount(item["credit"]),
                item["note"],
            ]
            for item in data.get("items", [])
        ]
        _write_sheet(ws, "科目明细账", filters, headers, rows)
        return wb, filters

    if report_key == "aux_ledger":
        data = get_aux_ledger(params)
        filters = {
            "book_id": data["book_id"],
            "aux_type": data["aux_type"],
            "subject_code": data["subject_code"],
            "aux_code": data["aux_code"],
            "start_date": data["start_date"],
            "end_date": data["end_date"],
        }
        headers = ["日期", "凭证号", "行号", "摘要", "科目", "辅助项目", "借方", "贷方", "备注"]
        rows = [
            [
                item["voucher_date"],
                f"{item['voucher_word']}{item['voucher_no']}",
                item["line_no"],
                item["summary"],
                f"{item['subject_code']} {item['subject_name']}",
                f"{item['aux_code']} {item['aux_name']}",
                _format_amount(item["debit"]),
                _format_amount(item["credit"]),
                item["note"],
            ]
            for item in data.get("items", [])
        ]
        _write_sheet(ws, "辅助明细账", filters, headers, rows)
        return wb, filters

    if report_key == "aux_balance":
        data = get_aux_balance(params)
        filters = {
            "book_id": data["book_id"],
            "aux_type": data["aux_type"],
            "subject_code": params.get("subject_code", ""),
            "aux_code": params.get("aux_code", ""),
            "start_date": data["start_date"],
            "end_date": data["end_date"],
            "primary": data["primary"],
        }
        headers = ["维度编码", "维度名称", "本期借方发生额", "本期贷方发生额", "期末余额"]
        rows = [
            [
                item["primary_code"],
                item["primary_name"],
                _format_amount(item["period_debit"]),
                _format_amount(item["period_credit"]),
                _format_amount(item["ending_balance"]),
            ]
            for item in data.get("items", [])
        ]
        _write_sheet(ws, "辅助余额表", filters, headers, rows)
        return wb, filters

    if report_key == "ar_ap_aging":
        data = get_aging_report(params)
        filters = {
            "book_id": data["book_id"],
            "start_date": data["start_date"],
            "end_date": data["end_date"],
            "as_of_date": data["as_of_date"],
        }
        headers = ["往来对象", "余额", "账期内金额", "逾期金额", "逾期天数"]
        rows = [
            [
                f"{item['counterparty_code']} {item['counterparty_name']}",
                _format_amount(item["balance"]),
                _format_amount(item["period_amount"]),
                _format_amount(item["overdue_amount"]),
                item["overdue_days"],
            ]
            for item in data.get("items", [])
        ]
        _write_sheet(ws, "账龄分析", filters, headers, rows)
        return wb, filters

    if report_key == "payments":
        data = list_payments(params)
        filters = {"book_id": data["book_id"], "status": params.get("status", "")}
        headers = ["ID", "标题", "收款人", "方式", "金额", "状态", "关联类型", "关联ID", "报销ID"]
        rows = [
            [
                item["id"],
                item["title"],
                item["payee_name"],
                item["pay_method"],
                _format_amount(item["amount"]),
                item["status"],
                item["related_type"],
                item["related_id"],
                item["reimbursement_id"],
            ]
            for item in data.get("items", [])
        ]
        _write_sheet(ws, "支付明细", filters, headers, rows)
        return wb, filters

    if report_key == "bank_reconcile":
        data = get_reconcile_report(params)
        filters = {"book_id": data["book_id"], "bank_account_id": data["bank_account_id"]}
        headers = ["指标", "值"]
        rows = [
            ["流水总笔数", data["total"]],
            ["未匹配笔数", data["unmatched_count"]],
            ["未匹配金额", _format_amount(data["unmatched_amount"])],
            ["已匹配金额", _format_amount(data["matched_amount"])],
            ["最新余额", _format_amount(data["latest_balance"]) if data["latest_balance"] is not None else ""],
        ]
        _write_sheet(ws, "银行对账报表", filters, headers, rows)
        return wb, filters

    if report_key == "tax_summary":
        data = get_tax_summary(params)
        filters = {"book_id": data["book_id"]}
        headers = ["类别", "金额", "税额"]
        rows = [
            [item["category"], _format_amount(item["total_amount"]), _format_amount(item["total_tax"])]
            for item in data.get("items", [])
        ]
        _write_sheet(ws, "税务汇总", filters, headers, rows)
        return wb, filters

    if report_key == "asset_ledger":
        data = get_asset_ledger(params)
        filters = {
            "book_id": data["book_id"],
            "asset_code": params.get("asset_code", ""),
            "asset_name": params.get("asset_name", ""),
            "category_id": params.get("category_id", ""),
            "status": params.get("status", ""),
            "department_id": params.get("department_id", ""),
            "person_id": params.get("person_id", ""),
            "start_use_from": params.get("start_use_from", ""),
            "start_use_to": params.get("start_use_to", ""),
            "dep_year": params.get("dep_year", ""),
            "dep_month": params.get("dep_month", ""),
        }
        headers = [
            "资产编号",
            "资产名称",
            "类别",
            "原值",
            "累计折旧",
            "净值",
            "状态",
            "部门",
            "责任人",
            "折旧方法",
            "使用年限(月)",
            "存放地点",
            "启用日期",
            "最近变动",
        ]
        rows = [
            [
                item["asset_code"],
                item["asset_name"],
                item["category_name"],
                _format_amount(item["original_value"]),
                _format_amount(item["accumulated_depr"]),
                _format_amount(item["net_value"]),
                item["status"],
                item["department_name"],
                item.get("person_name", ""),
                item.get("depreciation_method", ""),
                item.get("useful_life_months", ""),
                item.get("location", ""),
                item["start_use_date"],
                item.get("last_change_info", ""),
            ]
            for item in data.get("items", [])
        ]
        _write_sheet(ws, "固定资产台账", filters, headers, rows)
        return wb, filters

    if report_key == "asset_depreciation_detail":
        data = get_depreciation_detail(params)
        filters = {
            "book_id": data["book_id"],
            "year": data["period_year"],
            "month": data["period_month"],
            "asset_code": params.get("asset_code", ""),
            "asset_name": params.get("asset_name", ""),
            "category_id": params.get("category_id", ""),
        }
        headers = ["资产编号", "资产名称", "期间", "本期折旧金额", "累计折旧", "净值", "批次ID", "凭证ID"]
        rows = [
            [
                item["asset_code"],
                item["asset_name"],
                item["period"],
                _format_amount(item["amount"]),
                _format_amount(item.get("accumulated_depr", 0)),
                _format_amount(item.get("net_value", 0)),
                item["batch_id"],
                item["voucher_id"],
            ]
            for item in data.get("items", [])
        ]
        _write_sheet(ws, "折旧明细", filters, headers, rows)
        return wb, filters

    if report_key == "asset_depreciation_summary":
        data = get_depreciation_summary(params)
        filters = {
            "book_id": data["book_id"],
            "year": data["period_year"],
            "month": data["period_month"],
            "asset_code": params.get("asset_code", ""),
            "asset_name": params.get("asset_name", ""),
            "category_id": params.get("category_id", ""),
        }
        ws.title = "按类别"
        headers = ["类别", "本期折旧金额"]
        rows = [
            [item["category_name"], _format_amount(item["total_amount"])]
            for item in data.get("by_category", [])
        ]
        _write_sheet(ws, "折旧汇总-按类别", filters, headers, rows)

        ws2 = wb.create_sheet("按部门")
        headers2 = ["部门", "本期折旧金额"]
        rows2 = [
            [item["department_name"], _format_amount(item["total_amount"])]
            for item in data.get("by_department", [])
        ]
        _write_sheet(ws2, "折旧汇总-按部门", filters, headers2, rows2)
        return wb, filters

    raise ExportError("unsupported_report")


def export_report(report_key: str, params: Dict[str, str], operator: str = "") -> Tuple[bytes, str]:
    try:
        wb, filters = _build_workbook(report_key, params)
    except ExportError:
        raise
    except Exception as err:
        raise ExportError(str(err)) from err
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{report_key}_{timestamp}.xlsx"

    _audit_export(report_key, params.get("book_id") or filters.get("book_id"), filters, file_name, operator)

    return buffer.getvalue(), file_name
