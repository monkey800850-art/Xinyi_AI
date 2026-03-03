import io
import json
from datetime import date, datetime
from typing import Dict, List

from openpyxl import Workbook
from sqlalchemy import text

from app.db_router import get_connection_provider


class ConsolidationDisclosureAuditPackageError(RuntimeError):
    pass


RULE_CODE = "CONS30_DISCLOSURE_AUDIT_PACKAGE"


def _parse_positive_int(value: object, field: str) -> int:
    raw = str(value or "").strip()
    if not raw:
        raise ConsolidationDisclosureAuditPackageError(f"{field}_required")
    try:
        parsed = int(raw)
    except Exception as err:
        raise ConsolidationDisclosureAuditPackageError(f"{field}_invalid") from err
    if parsed <= 0:
        raise ConsolidationDisclosureAuditPackageError(f"{field}_invalid")
    return parsed


def _parse_period(value: object) -> str:
    raw = str(value or "").strip()
    if len(raw) != 7 or raw[4] != "-":
        raise ConsolidationDisclosureAuditPackageError("period_invalid")
    yy = raw[:4]
    mm = raw[5:]
    if not yy.isdigit() or not mm.isdigit():
        raise ConsolidationDisclosureAuditPackageError("period_invalid")
    month = int(mm)
    if month < 1 or month > 12:
        raise ConsolidationDisclosureAuditPackageError("period_invalid")
    return f"{int(yy):04d}-{month:02d}"


def _as_of_to_period(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        raise ConsolidationDisclosureAuditPackageError("as_of_required")
    try:
        parsed = date.fromisoformat(raw)
    except Exception as err:
        raise ConsolidationDisclosureAuditPackageError("as_of_invalid") from err
    return parsed.strftime("%Y-%m")


def _ensure_package_table(conn) -> None:
    conn.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS consolidation_audit_packages (
                id BIGINT PRIMARY KEY AUTO_INCREMENT,
                group_id BIGINT NOT NULL,
                period VARCHAR(7) NOT NULL,
                batch_id VARCHAR(64) NOT NULL,
                file_name VARCHAR(255) NOT NULL,
                source VARCHAR(32) NOT NULL DEFAULT 'generated',
                rule_code VARCHAR(64) NOT NULL DEFAULT 'CONS30_DISCLOSURE_AUDIT_PACKAGE',
                status VARCHAR(16) NOT NULL DEFAULT 'draft',
                package_meta_json JSON NOT NULL,
                package_blob LONGBLOB NOT NULL,
                operator_id BIGINT NOT NULL DEFAULT 0,
                created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_conso_audit_package (group_id, period)
            )
            """
        )
    )


def _build_workbook(group_id: int, period: str, reports: Dict[str, Dict[str, object]], flow: Dict[str, object]) -> bytes:
    wb = Workbook()
    ws_index = wb.active
    ws_index.title = "Disclosure_Index"

    ws_index["A1"] = "披露索引"
    ws_index["A2"] = "集团ID"
    ws_index["B2"] = group_id
    ws_index["A3"] = "期间"
    ws_index["B3"] = period
    ws_index["A4"] = "审批状态"
    ws_index["B4"] = str(flow.get("approval_status") or "")
    ws_index["A5"] = "索引条目数"
    ws_index["B5"] = "=COUNTA(Audit_Trail!A:A)-1"
    ws_index["A6"] = "校验时间"
    ws_index["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    report_sheets = [
        ("BALANCE_SHEET", "Balance_Sheet"),
        ("INCOME_STATEMENT", "Income_Statement"),
        ("CASH_FLOW", "Cash_Flow"),
        ("EQUITY_CHANGE", "Equity_Change"),
    ]

    for report_code, sheet_name in report_sheets:
        ws = wb.create_sheet(sheet_name)
        report = reports.get(report_code) or {}
        ws["A1"] = report_code
        ws["A2"] = "item_code"
        ws["B2"] = "label"
        ws["C2"] = "amount"
        for idx, item in enumerate((report.get("items") or []), start=3):
            ws[f"A{idx}"] = str(item.get("item_code") or "")
            ws[f"B{idx}"] = str(item.get("label") or "")
            ws[f"C{idx}"] = float(item.get("amount") or 0)
        ws["A20"] = "sheet_total"
        ws["C20"] = "=SUM(C3:C19)"

    ws_audit = wb.create_sheet("Audit_Trail")
    ws_audit.append(["index_id", "report_code", "item_code", "amount", "evidence_ref"])
    idx = 1
    for report_code, report in reports.items():
        for item in (report.get("items") or []):
            ws_audit.append(
                [
                    idx,
                    report_code,
                    str(item.get("item_code") or ""),
                    float(item.get("amount") or 0),
                    f"RPT-{group_id}-{period.replace('-', '')}-{report_code}",
                ]
            )
            idx += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_disclosure_and_audit_package(payload: Dict[str, object], operator_id: object = 1) -> Dict[str, object]:
    group_id = _parse_positive_int(payload.get("consolidation_group_id") or payload.get("group_id"), "consolidation_group_id")
    period_raw = str(payload.get("period") or "").strip()
    period = _parse_period(period_raw) if period_raw else _as_of_to_period(payload.get("as_of"))
    operator = _parse_positive_int(operator_id, "operator_id")
    batch_id = f"AUDPKG-{group_id}-{period.replace('-', '')}"
    file_name = f"consolidation_audit_package_{group_id}_{period.replace('-', '')}.xlsx"

    provider = get_connection_provider()
    with provider.begin() as conn:
        _ensure_package_table(conn)
        group_exists = conn.execute(text("SELECT id FROM consolidation_groups WHERE id=:gid LIMIT 1"), {"gid": group_id}).fetchone()
        if not group_exists:
            raise ConsolidationDisclosureAuditPackageError("consolidation_group_not_found")

        report_rows = conn.execute(
            text(
                """
                SELECT report_code, report_json
                FROM consolidation_report_snapshots
                WHERE group_id=:group_id AND period=:period
                ORDER BY report_code ASC
                """
            ),
            {"group_id": group_id, "period": period},
        ).fetchall()
        if len(report_rows) < 4:
            raise ConsolidationDisclosureAuditPackageError("report_snapshots_not_ready")
        reports: Dict[str, Dict[str, object]] = {}
        for row in report_rows:
            try:
                reports[str(row.report_code or "")] = json.loads(str(row.report_json or "{}"))
            except Exception:
                reports[str(row.report_code or "")] = {}
        required = {"BALANCE_SHEET", "INCOME_STATEMENT", "CASH_FLOW", "EQUITY_CHANGE"}
        missing = sorted(required - set(reports.keys()))
        if missing:
            raise ConsolidationDisclosureAuditPackageError(f"missing_reports:{','.join(missing)}")

        flow_row = conn.execute(
            text(
                """
                SELECT id, approval_status, check_result, check_note
                FROM consolidation_approval_flows
                WHERE group_id=:group_id AND period=:period
                ORDER BY id DESC
                LIMIT 1
                """
            ),
            {"group_id": group_id, "period": period},
        ).fetchone()
        if not flow_row:
            raise ConsolidationDisclosureAuditPackageError("approval_flow_not_found")
        if str(flow_row.check_result or "") != "passed":
            raise ConsolidationDisclosureAuditPackageError("approval_check_not_passed")

        flow = {
            "id": int(flow_row.id or 0),
            "approval_status": str(flow_row.approval_status or ""),
            "check_result": str(flow_row.check_result or ""),
            "check_note": str(flow_row.check_note or ""),
        }
        package_blob = _build_workbook(group_id, period, reports, flow)
        meta = {
            "group_id": group_id,
            "period": period,
            "report_count": 4,
            "flow_id": flow["id"],
            "approval_status": flow["approval_status"],
            "sheet_names": ["Disclosure_Index", "Balance_Sheet", "Income_Statement", "Cash_Flow", "Equity_Change", "Audit_Trail"],
        }

        existing = conn.execute(
            text(
                """
                SELECT id
                FROM consolidation_audit_packages
                WHERE group_id=:group_id AND period=:period
                LIMIT 1
                """
            ),
            {"group_id": group_id, "period": period},
        ).fetchone()
        if existing:
            conn.execute(
                text(
                    """
                    UPDATE consolidation_audit_packages
                    SET batch_id=:batch_id, file_name=:file_name, source='generated', rule_code=:rule_code,
                        status='draft', package_meta_json=:meta_json, package_blob=:package_blob,
                        operator_id=:operator_id, updated_at=NOW()
                    WHERE id=:id
                    """
                ),
                {
                    "id": int(existing.id),
                    "batch_id": batch_id,
                    "file_name": file_name,
                    "rule_code": RULE_CODE,
                    "meta_json": json.dumps(meta, ensure_ascii=False),
                    "package_blob": package_blob,
                    "operator_id": operator,
                },
            )
            package_id = int(existing.id)
        else:
            result = conn.execute(
                text(
                    """
                    INSERT INTO consolidation_audit_packages (
                        group_id, period, batch_id, file_name, source, rule_code, status,
                        package_meta_json, package_blob, operator_id
                    ) VALUES (
                        :group_id, :period, :batch_id, :file_name, 'generated', :rule_code, 'draft',
                        :meta_json, :package_blob, :operator_id
                    )
                    """
                ),
                {
                    "group_id": group_id,
                    "period": period,
                    "batch_id": batch_id,
                    "file_name": file_name,
                    "rule_code": RULE_CODE,
                    "meta_json": json.dumps(meta, ensure_ascii=False),
                    "package_blob": package_blob,
                    "operator_id": operator,
                },
            )
            package_id = int(result.lastrowid or 0)

    return {
        "group_id": group_id,
        "period": period,
        "rule_code": RULE_CODE,
        "package_id": package_id,
        "batch_id": batch_id,
        "file_name": file_name,
        "approval_status": flow["approval_status"],
        "sheet_names": meta["sheet_names"],
        "report_count": 4,
    }
