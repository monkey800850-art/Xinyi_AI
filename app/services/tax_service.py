import csv
import hashlib
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from sqlalchemy import text

from app.db import get_engine


class TaxError(RuntimeError):
    pass


BONUS_MONTHLY_BRACKETS = [
    (Decimal("3000"), Decimal("0.03"), Decimal("0")),
    (Decimal("12000"), Decimal("0.10"), Decimal("210")),
    (Decimal("25000"), Decimal("0.20"), Decimal("1410")),
    (Decimal("35000"), Decimal("0.25"), Decimal("2660")),
    (Decimal("55000"), Decimal("0.30"), Decimal("4410")),
    (Decimal("80000"), Decimal("0.35"), Decimal("7160")),
    (Decimal("999999999"), Decimal("0.45"), Decimal("15160")),
]

COMPREHENSIVE_ANNUAL_BRACKETS = [
    (Decimal("36000"), Decimal("0.03"), Decimal("0")),
    (Decimal("144000"), Decimal("0.10"), Decimal("2520")),
    (Decimal("300000"), Decimal("0.20"), Decimal("16920")),
    (Decimal("420000"), Decimal("0.25"), Decimal("31920")),
    (Decimal("660000"), Decimal("0.30"), Decimal("52920")),
    (Decimal("960000"), Decimal("0.35"), Decimal("85920")),
    (Decimal("999999999"), Decimal("0.45"), Decimal("181920")),
]

LABOR_BRACKETS = [
    (Decimal("20000"), Decimal("0.20"), Decimal("0")),
    (Decimal("50000"), Decimal("0.30"), Decimal("2000")),
    (Decimal("999999999"), Decimal("0.40"), Decimal("7000")),
]


INVOICE_HEADER_MAP = {
    "invoice_no": ["invoice_no", "发票号码", "发票号"],
    "invoice_code": ["invoice_code", "发票代码"],
    "invoice_date": ["invoice_date", "开票日期", "日期"],
    "amount": ["amount", "金额", "价税合计"],
    "tax_rate": ["tax_rate", "税率"],
    "tax_amount": ["tax_amount", "税额"],
    "seller_name": ["seller_name", "销方名称", "销售方", "销售方名称"],
    "buyer_name": ["buyer_name", "购方名称", "购买方", "购买方名称"],
    "category": ["category", "项目", "类别"],
}


def _normalize(s: str) -> str:
    return (s or "").strip().lower()


def _parse_date(value) -> date:
    if not value:
        return None
    if isinstance(value, date):
        return value
    value = str(value).strip()
    if "-" in value:
        return date.fromisoformat(value)
    if "/" in value:
        parts = value.split("/")
        return date(int(parts[0]), int(parts[1]), int(parts[2]))
    if len(value) == 8 and value.isdigit():
        return date(int(value[0:4]), int(value[4:6]), int(value[6:8]))
    raise ValueError("invalid_date")


def _parse_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        raise ValueError("invalid_amount")


def _pick_bracket(base: Decimal, brackets) -> Tuple[Decimal, Decimal]:
    for upper, rate, quick in brackets:
        if base <= upper:
            return rate, quick
    return brackets[-1][1], brackets[-1][2]


def calc_year_end_bonus_tax(payload: Dict[str, object]) -> Dict[str, object]:
    mode = (payload.get("tax_mode") or "").strip().lower()
    year = int(payload.get("biz_year") or 0) if payload.get("biz_year") not in (None, "") else None
    amount_raw = payload.get("bonus_amount")

    if amount_raw in (None, ""):
        raise TaxError("bonus_amount required")
    try:
        amount = _parse_decimal(amount_raw)
    except ValueError:
        raise TaxError("bonus_amount invalid")
    if amount <= 0:
        raise TaxError("bonus_amount must be > 0")
    if mode not in ("separate", "merge"):
        raise TaxError("tax_mode must be one of: separate, merge")

    if mode == "separate":
        taxable_base = (amount / Decimal("12")).quantize(Decimal("0.01"))
        rate, quick = _pick_bracket(taxable_base, BONUS_MONTHLY_BRACKETS)
        tax = (amount * rate - quick).quantize(Decimal("0.01"))
        explain = (
            "separate mode: bonus/12 choose monthly bracket; "
            "tax=bonus*rate-quick_deduction"
        )
    else:
        taxable_base = amount
        rate, quick = _pick_bracket(taxable_base, COMPREHENSIVE_ANNUAL_BRACKETS)
        tax = (taxable_base * rate - quick).quantize(Decimal("0.01"))
        explain = (
            "merge mode(simplified): treat bonus as annual comprehensive taxable base only; "
            "tax=base*rate-quick_deduction"
        )

    if tax < 0:
        tax = Decimal("0.00")

    return {
        "tax_mode": mode,
        "biz_year": year,
        "bonus_amount": float(amount),
        "taxable_base": float(taxable_base),
        "tax_amount": float(tax),
        "rate": float(rate),
        "quick_deduction": float(quick),
        "explain": explain,
    }


def calc_labor_service_tax(payload: Dict[str, object]) -> Dict[str, object]:
    gross_raw = payload.get("gross_amount")
    period = (payload.get("period") or "").strip() or None

    if gross_raw in (None, ""):
        raise TaxError("gross_amount required")
    try:
        gross = _parse_decimal(gross_raw)
    except ValueError:
        raise TaxError("gross_amount invalid")
    if gross <= 0:
        raise TaxError("gross_amount must be > 0")

    if gross <= Decimal("4000"):
        taxable_base = (gross - Decimal("800")).quantize(Decimal("0.01"))
        if taxable_base < 0:
            taxable_base = Decimal("0.00")
        base_explain = "gross<=4000, taxable_base=gross-800"
    else:
        taxable_base = (gross * Decimal("0.8")).quantize(Decimal("0.01"))
        base_explain = "gross>4000, taxable_base=gross*80%"

    rate, quick = _pick_bracket(taxable_base, LABOR_BRACKETS)
    tax = (taxable_base * rate - quick).quantize(Decimal("0.01"))
    if tax < 0:
        tax = Decimal("0.00")

    return {
        "period": period,
        "gross_amount": float(gross),
        "taxable_base": float(taxable_base),
        "tax_amount": float(tax),
        "rate": float(rate),
        "quick_deduction": float(quick),
        "explain": f"{base_explain}; labor bracket tax=base*rate-quick_deduction (simplified prewithholding)",
    }


def _detect_columns(headers: List[str], mapping: Dict[str, List[str]]) -> Dict[str, int]:
    index_map: Dict[str, int] = {}
    norm_headers = [_normalize(h) for h in headers]
    for key, aliases in mapping.items():
        for alias in aliases:
            alias_norm = _normalize(alias)
            if alias_norm in norm_headers:
                index_map[key] = norm_headers.index(alias_norm)
                break
    return index_map


def _rows_from_csv(file_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    text_data = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text_data))
    rows = list(reader)
    if not rows:
        raise TaxError("empty_file")
    return rows[0], rows[1:]


def _rows_from_excel(file_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise TaxError("empty_file")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = [list(r) for r in rows[1:]]
    return headers, data_rows


def _hash_invoice(book_id: int, invoice_code: str, invoice_no: str, invoice_date: date, amount: Decimal) -> str:
    raw = f"{book_id}|{invoice_code}|{invoice_no}|{invoice_date.isoformat()}|{amount}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _db_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return value


def _db_params(params: Dict[str, object]) -> Dict[str, object]:
    return {k: _db_value(v) for k, v in params.items()}


def _table_columns(conn, table_name: str) -> set[str]:
    try:
        rows = conn.execute(
            text(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = DATABASE()
                  AND table_name = :table_name
                """
            ),
            {"table_name": table_name},
        ).fetchall()
        return {str(r.column_name or "").strip().lower() for r in rows}
    except Exception:
        rows = conn.execute(text(f"PRAGMA table_info({table_name})")).fetchall()
        return {str(getattr(r, "name", r[1]) or "").strip().lower() for r in rows}


def create_tax_rule(payload: Dict[str, object]) -> Dict[str, object]:
    region = (payload.get("region") or "").strip()
    tax_type = (payload.get("tax_type") or "").strip()
    rate_raw = payload.get("rate")
    reduction_type = (payload.get("reduction_type") or "").strip() or None
    reduction_rate = payload.get("reduction_rate")
    note = (payload.get("note") or "").strip() or None
    is_enabled = int(payload.get("is_enabled", 1))

    if not region or not tax_type or rate_raw is None:
        raise TaxError("region/tax_type/rate required")

    try:
        rate = _parse_decimal(rate_raw)
    except ValueError:
        raise TaxError("rate invalid")

    rr = None
    if reduction_rate is not None and reduction_rate != "":
        rr = _parse_decimal(reduction_rate)

    engine = get_engine()
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                INSERT INTO tax_rules (region, tax_type, rate, reduction_type, reduction_rate, note, is_enabled)
                VALUES (:region, :tax_type, :rate, :reduction_type, :reduction_rate, :note, :is_enabled)
                """
            ),
            {
                "region": region,
                "tax_type": tax_type,
                "rate": rate,
                "reduction_type": reduction_type,
                "reduction_rate": rr,
                "note": note,
                "is_enabled": is_enabled,
            },
        )
        rule_id = result.lastrowid

    return {"id": rule_id}


def list_tax_rules() -> Dict[str, object]:
    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT id, region, tax_type, rate, reduction_type, reduction_rate, note, is_enabled
                FROM tax_rules
                ORDER BY id DESC
                """
            )
        ).fetchall()

    items = []
    for r in rows:
        items.append(
            {
                "id": r.id,
                "region": r.region,
                "tax_type": r.tax_type,
                "rate": float(r.rate),
                "reduction_type": r.reduction_type or "",
                "reduction_rate": float(r.reduction_rate) if r.reduction_rate is not None else None,
                "note": r.note or "",
                "is_enabled": int(r.is_enabled),
            }
        )

    return {"items": items}


def import_invoices(book_id: int, filename: str, file_bytes: bytes) -> Dict[str, object]:
    if not book_id:
        raise TaxError("book_id required")

    if filename.lower().endswith(".csv"):
        headers, data_rows = _rows_from_csv(file_bytes)
    elif filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        headers, data_rows = _rows_from_excel(file_bytes)
    else:
        raise TaxError("unsupported_file_type")

    index_map = _detect_columns(headers, INVOICE_HEADER_MAP)
    if "invoice_no" not in index_map or "invoice_date" not in index_map or "amount" not in index_map:
        raise TaxError("missing_required_columns: invoice_no, invoice_date, amount")

    total = 0
    success = 0
    failed = 0
    duplicated = 0
    errors: List[Dict[str, object]] = []

    engine = get_engine()
    with engine.begin() as conn:
        for i, row in enumerate(data_rows, start=2):
            if not any([str(c).strip() for c in row if c is not None]):
                continue
            total += 1
            try:
                invoice_no = str(row[index_map["invoice_no"]]).strip()
                invoice_code = (
                    str(row[index_map["invoice_code"]]).strip()
                    if "invoice_code" in index_map and row[index_map["invoice_code"]] is not None
                    else ""
                )
                invoice_date = _parse_date(row[index_map["invoice_date"]])
                amount = _parse_decimal(row[index_map["amount"]])
                tax_rate = (
                    _parse_decimal(row[index_map["tax_rate"]])
                    if "tax_rate" in index_map and row[index_map["tax_rate"]] not in (None, "")
                    else None
                )
                tax_amount = (
                    _parse_decimal(row[index_map["tax_amount"]])
                    if "tax_amount" in index_map and row[index_map["tax_amount"]] not in (None, "")
                    else None
                )
                seller_name = (
                    str(row[index_map["seller_name"]]).strip()
                    if "seller_name" in index_map and row[index_map["seller_name"]] is not None
                    else ""
                )
                buyer_name = (
                    str(row[index_map["buyer_name"]]).strip()
                    if "buyer_name" in index_map and row[index_map["buyer_name"]] is not None
                    else ""
                )
                category = (
                    str(row[index_map["category"]]).strip()
                    if "category" in index_map and row[index_map["category"]] is not None
                    else ""
                )

                if not invoice_no or not invoice_date:
                    raise ValueError("invalid_invoice")

                import_hash = _hash_invoice(book_id, invoice_code, invoice_no, invoice_date, amount)

                existing = conn.execute(
                    text("SELECT 1 FROM tax_invoices WHERE import_hash=:h"),
                    {"h": import_hash},
                ).fetchone()
                if existing:
                    duplicated += 1
                    continue

                conn.execute(
                    text(
                        """
                        INSERT INTO tax_invoices (
                            book_id, invoice_code, invoice_no, invoice_date,
                            amount, tax_rate, tax_amount, seller_name, buyer_name, category,
                            source_file, import_hash
                        ) VALUES (
                            :book_id, :invoice_code, :invoice_no, :invoice_date,
                            :amount, :tax_rate, :tax_amount, :seller_name, :buyer_name, :category,
                            :source_file, :import_hash
                        )
                        """
                    ),
                    {
                        "book_id": book_id,
                        "invoice_code": invoice_code,
                        "invoice_no": invoice_no,
                        "invoice_date": invoice_date,
                        "amount": amount,
                        "tax_rate": tax_rate,
                        "tax_amount": tax_amount,
                        "seller_name": seller_name,
                        "buyer_name": buyer_name,
                        "category": category,
                        "source_file": filename,
                        "import_hash": import_hash,
                    },
                )
                success += 1
            except Exception as err:
                failed += 1
                errors.append({"row": i, "error": str(err)})

    return {
        "total": total,
        "success": success,
        "failed": failed,
        "duplicated": duplicated,
        "errors": errors,
    }


def list_invoices(params: Dict[str, str]) -> Dict[str, object]:
    book_id_raw = (params.get("book_id") or "").strip()
    if not book_id_raw:
        raise TaxError("book_id required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")

    sql = """
        SELECT id, invoice_code, invoice_no, invoice_date, amount, tax_rate, tax_amount,
               seller_name, buyer_name, category
        FROM tax_invoices
        WHERE book_id=:book_id
        ORDER BY invoice_date DESC, id DESC
        LIMIT 200
    """

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(text(sql), {"book_id": book_id}).fetchall()

    items = []
    for r in rows:
        items.append(
            {
                "id": r.id,
                "invoice_code": r.invoice_code or "",
                "invoice_no": r.invoice_no,
                "invoice_date": r.invoice_date.isoformat(),
                "amount": float(r.amount),
                "tax_rate": float(r.tax_rate) if r.tax_rate is not None else None,
                "tax_amount": float(r.tax_amount) if r.tax_amount is not None else None,
                "seller_name": r.seller_name or "",
                "buyer_name": r.buyer_name or "",
                "category": r.category or "",
            }
        )

    return {"book_id": book_id, "items": items}


def get_tax_summary(params: Dict[str, str]) -> Dict[str, object]:
    book_id_raw = (params.get("book_id") or "").strip()
    if not book_id_raw:
        raise TaxError("book_id required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT category,
                       SUM(amount) AS total_amount,
                       SUM(tax_amount) AS total_tax
                FROM tax_invoices
                WHERE book_id=:book_id
                GROUP BY category
                ORDER BY category
                """
            ),
            {"book_id": book_id},
        ).fetchall()

    items = []
    for r in rows:
        items.append(
            {
                "category": r.category or "",
                "total_amount": float(Decimal(str(r.total_amount or 0))),
                "total_tax": float(Decimal(str(r.total_tax or 0))),
            }
        )

    return {"book_id": book_id, "items": items}


def validate_tax(params: Dict[str, str]) -> Dict[str, object]:
    book_id_raw = (params.get("book_id") or "").strip()
    if not book_id_raw:
        raise TaxError("book_id required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")

    errors: List[Dict[str, object]] = []

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT id, invoice_no, invoice_date, amount, tax_rate, tax_amount, seller_name, buyer_name
                FROM tax_invoices
                WHERE book_id=:book_id
                """
            ),
            {"book_id": book_id},
        ).fetchall()

    for r in rows:
        if not r.seller_name or not r.buyer_name:
            errors.append({"id": r.id, "message": "缺少购销方名称"})
        if r.tax_rate is not None and r.tax_amount is not None:
            expected = (Decimal(str(r.amount)) * Decimal(str(r.tax_rate))).quantize(Decimal("0.01"))
            actual = Decimal(str(r.tax_amount)).quantize(Decimal("0.01"))
            if expected != actual:
                errors.append({"id": r.id, "message": "税额与税率不一致"})

    return {"book_id": book_id, "errors": errors}


def build_tax_alerts(params: Dict[str, str]) -> Dict[str, object]:
    book_id_raw = (params.get("book_id") or "").strip()
    if not book_id_raw:
        raise TaxError("book_id required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")

    alerts: List[Dict[str, object]] = []

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT id, invoice_no, amount, tax_rate, tax_amount, seller_name, buyer_name
                FROM tax_invoices
                WHERE book_id=:book_id
                """
            ),
            {"book_id": book_id},
        ).fetchall()

    for r in rows:
        if not r.seller_name or not r.buyer_name:
            alerts.append({"alert_type": "missing_party", "severity": "warning", "message": f"发票{r.invoice_no}缺少购销方", "ref_id": r.id})
        if r.tax_rate is None:
            alerts.append({"alert_type": "missing_rate", "severity": "warning", "message": f"发票{r.invoice_no}缺少税率", "ref_id": r.id})
        if r.tax_rate is not None and r.tax_amount is not None:
            expected = (Decimal(str(r.amount)) * Decimal(str(r.tax_rate))).quantize(Decimal("0.01"))
            actual = Decimal(str(r.tax_amount)).quantize(Decimal("0.01"))
            if expected != actual:
                alerts.append({"alert_type": "tax_mismatch", "severity": "risk", "message": f"发票{r.invoice_no}税额不一致", "ref_id": r.id})

    # persist alerts
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM tax_alerts WHERE book_id=:book_id"), {"book_id": book_id})
        for a in alerts:
            conn.execute(
                text(
                    """
                    INSERT INTO tax_alerts (book_id, alert_type, severity, message, ref_type, ref_id)
                    VALUES (:book_id, :alert_type, :severity, :message, 'invoice', :ref_id)
                    """
                ),
                {"book_id": book_id, **a},
            )

    return {"book_id": book_id, "items": alerts}


def list_tax_alerts(params: Dict[str, str]) -> Dict[str, object]:
    book_id_raw = (params.get("book_id") or "").strip()
    if not book_id_raw:
        raise TaxError("book_id required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT id, alert_type, severity, message, ref_type, ref_id, created_at
                FROM tax_alerts
                WHERE book_id=:book_id
                ORDER BY id DESC
                """
            ),
            {"book_id": book_id},
        ).fetchall()

    items = []
    for r in rows:
        items.append(
            {
                "id": r.id,
                "alert_type": r.alert_type,
                "severity": r.severity,
                "message": r.message,
                "ref_type": r.ref_type or "",
                "ref_id": r.ref_id or "",
                "created_at": r.created_at.isoformat(),
            }
        )

    return {"book_id": book_id, "items": items}


def validate_invoice(invoice: Dict[str, object]) -> Dict[str, object]:
    invoice_no = str(invoice.get("invoice_no") or "").strip()
    invoice_code = str(invoice.get("invoice_code") or "").strip()
    seller_name = str(invoice.get("seller_name") or "").strip()
    buyer_name = str(invoice.get("buyer_name") or "").strip()
    errors: List[Dict[str, str]] = []

    if not invoice_no:
        errors.append({"field": "invoice_no", "message": "发票号码不能为空"})
    if not invoice_code:
        errors.append({"field": "invoice_code", "message": "发票代码不能为空"})
    if not seller_name:
        errors.append({"field": "seller_name", "message": "销方名称不能为空"})
    if not buyer_name:
        errors.append({"field": "buyer_name", "message": "购方名称不能为空"})

    invoice_date_raw = invoice.get("invoice_date")
    parsed_date = None
    try:
        parsed_date = _parse_date(invoice_date_raw)
    except Exception:
        errors.append({"field": "invoice_date", "message": "开票日期非法"})
    if parsed_date and parsed_date > date.today():
        errors.append({"field": "invoice_date", "message": "开票日期不能晚于今天"})

    amount = Decimal("0")
    try:
        amount = _parse_decimal(invoice.get("amount"))
        if amount <= 0:
            errors.append({"field": "amount", "message": "金额必须大于0"})
    except Exception:
        errors.append({"field": "amount", "message": "金额格式非法"})

    tax_rate = None
    tax_amount = None
    try:
        if invoice.get("tax_rate") not in (None, ""):
            tax_rate = _parse_decimal(invoice.get("tax_rate"))
    except Exception:
        errors.append({"field": "tax_rate", "message": "税率格式非法"})
    try:
        if invoice.get("tax_amount") not in (None, ""):
            tax_amount = _parse_decimal(invoice.get("tax_amount"))
    except Exception:
        errors.append({"field": "tax_amount", "message": "税额格式非法"})

    if tax_rate is not None and tax_amount is not None and amount > 0:
        expected = (amount * tax_rate).quantize(Decimal("0.01"))
        actual = tax_amount.quantize(Decimal("0.01"))
        if expected != actual:
            errors.append({"field": "tax_amount", "message": "税额与税率计算不一致"})

    return {"valid": len(errors) == 0, "errors": errors}


def verify_invoice(payload: Dict[str, object]) -> Dict[str, object]:
    invoice_id = payload.get("invoice_id")
    engine = get_engine()

    invoice = dict(payload)
    if invoice_id not in (None, ""):
        try:
            iid = int(invoice_id)
        except Exception:
            raise TaxError("invoice_id invalid")
        with engine.begin() as conn:
            row = conn.execute(
                text(
                    """
                    SELECT id, invoice_code, invoice_no, invoice_date, amount, tax_rate, tax_amount, seller_name, buyer_name
                    FROM tax_invoices
                    WHERE id=:id
                    """
                ),
                {"id": iid},
            ).fetchone()
            if not row:
                raise TaxError("invoice_not_found")
            invoice = {
                "invoice_code": row.invoice_code,
                "invoice_no": row.invoice_no,
                "invoice_date": row.invoice_date,
                "amount": row.amount,
                "tax_rate": row.tax_rate,
                "tax_amount": row.tax_amount,
                "seller_name": row.seller_name,
                "buyer_name": row.buyer_name,
            }
            result = validate_invoice(invoice)
            cols = _table_columns(conn, "tax_invoices")
            if {"verification_status", "verification_message", "verified_at"}.issubset(cols):
                conn.execute(
                    text(
                        """
                        UPDATE tax_invoices
                        SET verification_status=:verification_status,
                            verification_message=:verification_message,
                            verified_at=:verified_at
                        WHERE id=:id
                        """
                    ),
                    _db_params(
                        {
                            "id": iid,
                            "verification_status": "passed" if result["valid"] else "failed",
                            "verification_message": "" if result["valid"] else "; ".join([e["message"] for e in result["errors"]]),
                            "verified_at": datetime.now(),
                        }
                    ),
                )
            return {"invoice_id": iid, **result}

    return validate_invoice(invoice)


def create_tax_diff_entry(payload: Dict[str, object]) -> Dict[str, object]:
    book_id_raw = payload.get("book_id")
    period = str(payload.get("period") or "").strip()
    tax_subject = str(payload.get("tax_subject") or "").strip()
    reason_code = str(payload.get("reason_code") or "").strip() or "MANUAL"
    remark = str(payload.get("remark") or "").strip()

    if book_id_raw in (None, ""):
        raise TaxError("book_id required")
    if not period:
        raise TaxError("period required")
    if not tax_subject:
        raise TaxError("tax_subject required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")
    try:
        tax_amount = _parse_decimal(payload.get("tax_amount"))
        accounting_amount = _parse_decimal(payload.get("accounting_amount"))
    except Exception:
        raise TaxError("tax_amount/accounting_amount invalid")
    diff_amount = (tax_amount - accounting_amount).quantize(Decimal("0.01"))

    engine = get_engine()
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                INSERT INTO tax_difference_ledger (
                    book_id, period, tax_subject, tax_amount, accounting_amount, diff_amount,
                    reason_code, remark, created_at
                ) VALUES (
                    :book_id, :period, :tax_subject, :tax_amount, :accounting_amount, :diff_amount,
                    :reason_code, :remark, :created_at
                )
                """
            ),
            _db_params(
                {
                    "book_id": book_id,
                    "period": period,
                    "tax_subject": tax_subject,
                    "tax_amount": tax_amount,
                    "accounting_amount": accounting_amount,
                    "diff_amount": diff_amount,
                    "reason_code": reason_code,
                    "remark": remark,
                    "created_at": datetime.now(),
                }
            ),
        )
        entry_id = int(result.lastrowid)
    return {"id": entry_id, "book_id": book_id, "period": period, "diff_amount": float(diff_amount)}


def list_tax_diff_entries(params: Dict[str, str]) -> Dict[str, object]:
    book_id_raw = (params.get("book_id") or "").strip()
    period = (params.get("period") or "").strip()
    if not book_id_raw:
        raise TaxError("book_id required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")

    sql = """
        SELECT id, period, tax_subject, tax_amount, accounting_amount, diff_amount, reason_code, remark, created_at
        FROM tax_difference_ledger
        WHERE book_id=:book_id
    """
    args = {"book_id": book_id}
    if period:
        sql += " AND period=:period"
        args["period"] = period
    sql += " ORDER BY id DESC LIMIT 300"

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(text(sql), args).fetchall()
    return {
        "book_id": book_id,
        "items": [
            {
                "id": int(r.id),
                "period": str(r.period),
                "tax_subject": str(r.tax_subject or ""),
                "tax_amount": float(r.tax_amount or 0),
                "accounting_amount": float(r.accounting_amount or 0),
                "diff_amount": float(r.diff_amount or 0),
                "reason_code": str(r.reason_code or ""),
                "remark": str(r.remark or ""),
                "created_at": str(r.created_at or ""),
            }
            for r in rows
        ],
    }


def map_tax_declaration(payload: Dict[str, object]) -> Dict[str, object]:
    book_id_raw = payload.get("book_id")
    declaration_code = str(payload.get("declaration_code") or "").strip()
    mappings = payload.get("mappings")
    if book_id_raw in (None, ""):
        raise TaxError("book_id required")
    if not declaration_code:
        raise TaxError("declaration_code required")
    if not isinstance(mappings, list) or not mappings:
        raise TaxError("mappings required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")

    normalized: List[Dict[str, str]] = []
    for idx, m in enumerate(mappings):
        if not isinstance(m, dict):
            raise TaxError(f"mapping[{idx}] invalid")
        worksheet_cell = str(m.get("worksheet_cell") or "").strip()
        source_type = str(m.get("source_type") or "").strip()
        source_key = str(m.get("source_key") or "").strip()
        expression = str(m.get("expression") or "").strip()
        if not worksheet_cell or not source_type or not source_key:
            raise TaxError(f"mapping[{idx}] missing required fields")
        normalized.append(
            {
                "worksheet_cell": worksheet_cell,
                "source_type": source_type,
                "source_key": source_key,
                "expression": expression,
            }
        )

    engine = get_engine()
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                UPDATE tax_declaration_mappings
                SET is_enabled=0
                WHERE book_id=:book_id AND declaration_code=:declaration_code
                """
            ),
            {"book_id": book_id, "declaration_code": declaration_code},
        )
        for row in normalized:
            conn.execute(
                text(
                    """
                    INSERT INTO tax_declaration_mappings (
                        book_id, declaration_code, worksheet_cell, source_type, source_key, expression, is_enabled, created_at
                    ) VALUES (
                        :book_id, :declaration_code, :worksheet_cell, :source_type, :source_key, :expression, 1, :created_at
                    )
                    """
                ),
                _db_params(
                    {
                        "book_id": book_id,
                        "declaration_code": declaration_code,
                        **row,
                        "created_at": datetime.now(),
                    }
                ),
            )
    return {"book_id": book_id, "declaration_code": declaration_code, "count": len(normalized)}


def list_tax_declaration_mappings(params: Dict[str, str]) -> Dict[str, object]:
    book_id_raw = (params.get("book_id") or "").strip()
    declaration_code = (params.get("declaration_code") or "").strip()
    if not book_id_raw:
        raise TaxError("book_id required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")
    if not declaration_code:
        raise TaxError("declaration_code required")

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT id, worksheet_cell, source_type, source_key, expression, is_enabled
                FROM tax_declaration_mappings
                WHERE book_id=:book_id
                  AND declaration_code=:declaration_code
                  AND is_enabled=1
                ORDER BY worksheet_cell ASC
                """
            ),
            {"book_id": book_id, "declaration_code": declaration_code},
        ).fetchall()
    return {
        "book_id": book_id,
        "declaration_code": declaration_code,
        "items": [
            {
                "id": int(r.id),
                "worksheet_cell": str(r.worksheet_cell),
                "source_type": str(r.source_type),
                "source_key": str(r.source_key),
                "expression": str(r.expression or ""),
                "is_enabled": int(r.is_enabled or 0),
            }
            for r in rows
        ],
    }


def build_tax_declaration_mapping(payload: Dict[str, object]) -> Dict[str, object]:
    # Resolve declaration mapping with provided source values.
    book_id_raw = payload.get("book_id")
    declaration_code = str(payload.get("declaration_code") or "").strip()
    source_values = payload.get("source_values") if isinstance(payload.get("source_values"), dict) else {}
    if book_id_raw in (None, ""):
        raise TaxError("book_id required")
    if not declaration_code:
        raise TaxError("declaration_code required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise TaxError("book_id must be integer")

    mapping = list_tax_declaration_mappings({"book_id": str(book_id), "declaration_code": declaration_code})
    cells: List[Dict[str, object]] = []
    for m in mapping["items"]:
        source_key = str(m["source_key"])
        value = source_values.get(source_key)
        if value is None:
            value = 0
        try:
            numeric = _parse_decimal(value)
        except Exception:
            numeric = Decimal("0")
        cells.append(
            {
                "worksheet_cell": str(m["worksheet_cell"]),
                "source_key": source_key,
                "value": float(numeric),
                "expression": str(m.get("expression") or ""),
            }
        )
    return {"book_id": book_id, "declaration_code": declaration_code, "cells": cells}
