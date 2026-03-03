import csv
import hashlib
import io
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from sqlalchemy import text

from app.db import get_engine


class BankImportError(RuntimeError):
    pass


HEADER_MAP = {
    "date": ["date", "交易日期", "日期", "记账日期"],
    "amount": ["amount", "交易金额", "金额"],
    "summary": ["summary", "摘要", "用途", "说明"],
    "counterparty": ["counterparty", "对方户名", "对手户名", "对方名称"],
    "balance": ["balance", "余额"],
    "serial_no": ["serial_no", "流水号", "交易流水号", "序号"],
}
REQUIRED_FIELDS = ("date", "amount")
SUPPORTED_FIELDS = ("date", "amount", "summary", "counterparty", "balance", "serial_no")


def _normalize(s: str) -> str:
    return (s or "").strip().lower()


def _parse_date(value: str) -> date:
    if not value:
        return None
    if isinstance(value, date):
        return value
    value = str(value).strip()
    if "-" in value:
        return date.fromisoformat(value)
    if "/" in value:
        parts = value.split("/")
        return date(int(parts[0]), int(parts[1]), int(parts[2]))
    if len(value) == 8 and value.isdigit():
        return date(int(value[0:4]), int(value[4:6]), int(value[6:8]))
    raise ValueError("invalid_date")


def _parse_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        raise ValueError("invalid_amount")


def _detect_columns(headers: List[str]) -> Dict[str, int]:
    index_map: Dict[str, int] = {}
    norm_headers = [_normalize(h) for h in headers]

    for key, aliases in HEADER_MAP.items():
        for alias in aliases:
            alias_norm = _normalize(alias)
            if alias_norm in norm_headers:
                index_map[key] = norm_headers.index(alias_norm)
                break

    return index_map


def map_template(headers: List[str], template_mapping: Dict[str, object]) -> Dict[str, int]:
    index_map: Dict[str, int] = {}
    norm_headers = [_normalize(h) for h in headers]
    for field, source in (template_mapping or {}).items():
        key = _normalize(str(field))
        if key not in SUPPORTED_FIELDS:
            continue
        if isinstance(source, int):
            if 0 <= source < len(headers):
                index_map[key] = int(source)
            continue
        raw = str(source or "").strip()
        if raw.isdigit():
            idx = int(raw)
            if 0 <= idx < len(headers):
                index_map[key] = idx
            continue
        source_norm = _normalize(raw)
        if source_norm in norm_headers:
            index_map[key] = norm_headers.index(source_norm)
    return index_map


def _hash_row(book_id: int, bank_account_id: int, txn_date: date, amount: Decimal, summary: str, serial_no: str) -> str:
    raw = f"{book_id}|{bank_account_id}|{txn_date.isoformat()}|{amount}|{summary}|{serial_no}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _db_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date):
        return value.isoformat()
    return value


def _db_params(params: Dict[str, object]) -> Dict[str, object]:
    return {k: _db_value(v) for k, v in params.items()}


def _insert_row(conn, row: Dict[str, object]) -> bool:
    existing = conn.execute(
        text("SELECT 1 FROM bank_transactions WHERE import_hash=:h"),
        {"h": row["import_hash"]},
    ).fetchone()
    if existing:
        return False

    conn.execute(
        text(
            """
            INSERT INTO bank_transactions (
                book_id, bank_account_id, txn_date, amount, summary, counterparty,
                balance, serial_no, currency, source_file, import_hash
            ) VALUES (
                :book_id, :bank_account_id, :txn_date, :amount, :summary, :counterparty,
                :balance, :serial_no, :currency, :source_file, :import_hash
            )
            """
        ),
        _db_params(row),
    )
    return True


def _rows_from_csv(file_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    text_data = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text_data))
    rows = list(reader)
    if not rows:
        raise BankImportError("empty_file")
    headers = rows[0]
    data_rows = rows[1:]
    return headers, data_rows


def _rows_from_excel(file_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise BankImportError("empty_file")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = [list(r) for r in rows[1:]]
    return headers, data_rows


def _classify_row_error(err: Exception) -> str:
    msg = str(err or "").lower()
    if "invalid_date" in msg:
        return "invalid_date"
    if "invalid_amount" in msg:
        return "invalid_amount"
    if "index" in msg or "list index out of range" in msg:
        return "column_mapping_error"
    return "row_parse_error"


def handle_import_error(err: Exception, row_no: int, row_data: List[object]) -> Dict[str, object]:
    return {
        "row": int(row_no),
        "error_code": _classify_row_error(err),
        "message": str(err),
        "row_data": ["" if c is None else str(c) for c in (row_data or [])],
    }


def import_bank_transactions(
    book_id: int,
    bank_account_id: int,
    filename: str,
    file_bytes: bytes,
    template_mapping: Dict[str, object] | None = None,
) -> Dict[str, object]:
    if not book_id or not bank_account_id:
        raise BankImportError("book_id and bank_account_id required")

    if filename.lower().endswith(".csv"):
        headers, data_rows = _rows_from_csv(file_bytes)
    elif filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        headers, data_rows = _rows_from_excel(file_bytes)
    else:
        raise BankImportError("unsupported_file_type")

    # Template mapping takes precedence; unmapped fields fall back to auto-detection.
    index_map = map_template(headers, template_mapping or {})
    auto_map = _detect_columns(headers)
    for field, idx in auto_map.items():
        if field not in index_map:
            index_map[field] = idx
    if "date" not in index_map or "amount" not in index_map:
        raise BankImportError("missing_required_columns: date, amount")

    total = 0
    success = 0
    failed = 0
    duplicated = 0
    seen_in_file = set()
    errors: List[Dict[str, object]] = []
    receipts: List[Dict[str, object]] = []

    engine = get_engine()
    with engine.begin() as conn:
        for i, row in enumerate(data_rows, start=2):
            if not any([str(c).strip() for c in row if c is not None]):
                continue
            total += 1
            try:
                txn_date = _parse_date(row[index_map["date"]])
                amount = _parse_decimal(row[index_map["amount"]])
                summary = (
                    str(row[index_map["summary"]]).strip()
                    if "summary" in index_map and row[index_map["summary"]] is not None
                    else ""
                )
                counterparty = (
                    str(row[index_map["counterparty"]]).strip()
                    if "counterparty" in index_map and row[index_map["counterparty"]] is not None
                    else ""
                )
                balance = (
                    _parse_decimal(row[index_map["balance"]])
                    if "balance" in index_map and row[index_map["balance"]] not in (None, "")
                    else None
                )
                serial_no = (
                    str(row[index_map["serial_no"]]).strip()
                    if "serial_no" in index_map and row[index_map["serial_no"]] is not None
                    else ""
                )

                if not txn_date:
                    raise ValueError("invalid_date")

                import_hash = _hash_row(book_id, bank_account_id, txn_date, amount, summary, serial_no)
                if import_hash in seen_in_file:
                    duplicated += 1
                    continue
                seen_in_file.add(import_hash)
                inserted = _insert_row(
                    conn,
                    {
                        "book_id": book_id,
                        "bank_account_id": bank_account_id,
                        "txn_date": txn_date,
                        "amount": amount,
                        "summary": summary,
                        "counterparty": counterparty,
                        "balance": balance,
                        "serial_no": serial_no,
                        "currency": "CNY",
                        "source_file": filename,
                        "import_hash": import_hash,
                    },
                )
                if inserted:
                    success += 1
                else:
                    duplicated += 1
            except Exception as err:
                failed += 1
                receipt = handle_import_error(err, i, row)
                receipts.append(receipt)
                errors.append({"row": i, "error": str(err)})

    return {
        "total": total,
        "success": success,
        "failed": failed,
        "duplicated": duplicated,
        "errors": errors,
        "template_mapping_used": {k: headers[v] for k, v in index_map.items() if 0 <= int(v) < len(headers)},
        "exception_receipt": {
            "status": "ok" if failed == 0 else "partial_failed",
            "error_count": failed,
            "errors": receipts,
        },
    }


def list_bank_transactions(params: Dict[str, str]) -> Dict[str, object]:
    book_id_raw = (params.get("book_id") or "").strip()
    bank_account_raw = (params.get("bank_account_id") or "").strip()
    if not book_id_raw:
        raise BankImportError("book_id required")
    try:
        book_id = int(book_id_raw)
    except Exception:
        raise BankImportError("book_id must be integer")

    bank_account_id = None
    if bank_account_raw:
        try:
            bank_account_id = int(bank_account_raw)
        except Exception:
            raise BankImportError("bank_account_id must be integer")

    sql = """
        SELECT id, bank_account_id, txn_date, amount, summary, counterparty, balance, serial_no
        FROM bank_transactions
        WHERE book_id=:book_id
    """
    params_sql = {"book_id": book_id}
    if bank_account_id:
        sql += " AND bank_account_id=:bank_account_id"
        params_sql["bank_account_id"] = bank_account_id
    sql += " ORDER BY txn_date DESC, id DESC LIMIT 200"

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(text(sql), params_sql).fetchall()

    items = []
    for r in rows:
        items.append(
            {
                "id": r.id,
                "bank_account_id": r.bank_account_id,
                "txn_date": r.txn_date.isoformat(),
                "amount": float(r.amount),
                "summary": r.summary or "",
                "counterparty": r.counterparty or "",
                "balance": float(r.balance) if r.balance is not None else None,
                "serial_no": r.serial_no or "",
            }
        )

    return {"book_id": book_id, "items": items}
