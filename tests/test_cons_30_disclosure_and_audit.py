import io
import json
import os
import runpy
import sys
import time
import unittest
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app.db import get_engine


class Cons30DisclosureAuditPackageTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        os.environ["DB_HOST"] = "127.0.0.1"
        os.environ["DB_PORT"] = "3306"
        os.environ["DB_NAME"] = "xinyi_ai"
        os.environ["DB_USER"] = "root"
        os.environ["DB_PASSWORD"] = "88888888"

        repo_root = Path(__file__).resolve().parents[1]
        sys.path.insert(0, str(repo_root))
        ns = runpy.run_path(str(repo_root / "app.py"))
        cls.app = ns["create_app"]()
        cls.client = cls.app.test_client()
        cls.engine = get_engine()
        cls.sid = str(int(time.time()))[-6:]
        cls._ensure_tables()

    @classmethod
    def _ensure_tables(cls):
        with cls.engine.begin() as conn:
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS consolidation_groups (
                        id BIGINT PRIMARY KEY AUTO_INCREMENT,
                        group_code VARCHAR(64) NOT NULL UNIQUE,
                        group_name VARCHAR(128) NOT NULL,
                        group_type VARCHAR(32) NOT NULL DEFAULT 'standard',
                        status VARCHAR(16) NOT NULL DEFAULT 'active',
                        is_enabled TINYINT NOT NULL DEFAULT 1,
                        note VARCHAR(255) NULL,
                        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS consolidation_adjustments (
                        id BIGINT PRIMARY KEY AUTO_INCREMENT,
                        group_id BIGINT NOT NULL,
                        period VARCHAR(7) NOT NULL,
                        status VARCHAR(16) NOT NULL DEFAULT 'active',
                        operator_id BIGINT NOT NULL,
                        lines_json JSON NOT NULL,
                        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS consolidation_report_snapshots (
                        id BIGINT PRIMARY KEY AUTO_INCREMENT,
                        group_id BIGINT NOT NULL,
                        period VARCHAR(7) NOT NULL,
                        report_code VARCHAR(64) NOT NULL,
                        template_code VARCHAR(64) NOT NULL,
                        source VARCHAR(32) NOT NULL DEFAULT 'generated',
                        rule_code VARCHAR(64) NOT NULL DEFAULT 'CONS_REPORTS_GEN',
                        batch_id VARCHAR(64) NOT NULL,
                        status VARCHAR(16) NOT NULL DEFAULT 'draft',
                        template_json JSON NOT NULL,
                        report_json JSON NOT NULL,
                        operator_id BIGINT NOT NULL DEFAULT 0,
                        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        UNIQUE KEY uq_conso_report_snapshot (group_id, period, report_code)
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS consolidation_approval_flows (
                        id BIGINT PRIMARY KEY AUTO_INCREMENT,
                        group_id BIGINT NOT NULL,
                        period VARCHAR(7) NOT NULL,
                        batch_id VARCHAR(64) NOT NULL,
                        check_result VARCHAR(16) NOT NULL DEFAULT 'failed',
                        approval_status VARCHAR(16) NOT NULL DEFAULT 'submitted',
                        approver_id BIGINT NULL,
                        operator_id BIGINT NOT NULL DEFAULT 0,
                        check_note VARCHAR(255) NULL,
                        check_payload_json JSON NULL,
                        approved_at DATETIME NULL,
                        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        UNIQUE KEY uq_conso_approval_flow (group_id, period)
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS consolidation_audit_packages (
                        id BIGINT PRIMARY KEY AUTO_INCREMENT,
                        group_id BIGINT NOT NULL,
                        period VARCHAR(7) NOT NULL,
                        batch_id VARCHAR(64) NOT NULL,
                        file_name VARCHAR(255) NOT NULL,
                        source VARCHAR(32) NOT NULL DEFAULT 'generated',
                        rule_code VARCHAR(64) NOT NULL DEFAULT 'CONS30_DISCLOSURE_AUDIT_PACKAGE',
                        status VARCHAR(16) NOT NULL DEFAULT 'draft',
                        package_meta_json JSON NOT NULL,
                        package_blob LONGBLOB NOT NULL,
                        operator_id BIGINT NOT NULL DEFAULT 0,
                        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        UNIQUE KEY uq_conso_audit_package (group_id, period)
                    )
                    """
                )
            )

            def _has_column(name: str) -> bool:
                row = conn.execute(
                    text(
                        """
                        SELECT COUNT(*) AS cnt
                        FROM information_schema.columns
                        WHERE table_schema = DATABASE()
                          AND table_name='consolidation_adjustments'
                          AND column_name=:name
                        """
                    ),
                    {"name": name},
                ).fetchone()
                return int(row.cnt or 0) > 0

            optional_columns = [
                ("source", "ALTER TABLE consolidation_adjustments ADD COLUMN source VARCHAR(32) NULL"),
                ("rule_code", "ALTER TABLE consolidation_adjustments ADD COLUMN rule_code VARCHAR(64) NULL"),
                ("evidence_ref", "ALTER TABLE consolidation_adjustments ADD COLUMN evidence_ref VARCHAR(255) NULL"),
                ("batch_id", "ALTER TABLE consolidation_adjustments ADD COLUMN batch_id VARCHAR(64) NULL"),
                ("tag", "ALTER TABLE consolidation_adjustments ADD COLUMN tag VARCHAR(64) NULL"),
                ("reviewed_by", "ALTER TABLE consolidation_adjustments ADD COLUMN reviewed_by BIGINT NULL"),
                ("reviewed_at", "ALTER TABLE consolidation_adjustments ADD COLUMN reviewed_at DATETIME NULL"),
                ("locked_by", "ALTER TABLE consolidation_adjustments ADD COLUMN locked_by BIGINT NULL"),
                ("locked_at", "ALTER TABLE consolidation_adjustments ADD COLUMN locked_at DATETIME NULL"),
                ("note", "ALTER TABLE consolidation_adjustments ADD COLUMN note VARCHAR(255) NULL"),
            ]
            for name, ddl in optional_columns:
                if not _has_column(name):
                    conn.execute(text(ddl))

    def _create_group(self) -> int:
        resp = self.client.post(
            "/api/consolidation/groups",
            json={
                "group_code": f"CONS30-{self.sid}-{self._testMethodName[-2:]}",
                "group_name": f"CONS30组{self.sid}",
            },
        )
        self.assertEqual(resp.status_code, 201, resp.get_data(as_text=True))
        return int(resp.get_json()["id"])

    def _insert_source_adjustment(self, gid: int):
        lines = [
            {"subject_code": "1001", "debit": "100.00", "credit": "20.00", "note": "cash", "set_id": "SRC-30", "source": "generated", "rule": "UP_INV", "evidence_ref": "SRC-30", "operator_id": "1"},
            {"subject_code": "6001", "debit": "0", "credit": "50.00", "note": "revenue", "set_id": "SRC-30", "source": "generated", "rule": "UP_INV", "evidence_ref": "SRC-30", "operator_id": "1"},
            {"subject_code": "5001", "debit": "30.00", "credit": "0", "note": "cost", "set_id": "SRC-30", "source": "generated", "rule": "UP_INV", "evidence_ref": "SRC-30", "operator_id": "1"},
            {"subject_code": "6601", "debit": "10.00", "credit": "0", "note": "expense", "set_id": "SRC-30", "source": "generated", "rule": "UP_INV", "evidence_ref": "SRC-30", "operator_id": "1"},
            {"subject_code": "2201", "debit": "0", "credit": "40.00", "note": "liability", "set_id": "SRC-30", "source": "generated", "rule": "UP_INV", "evidence_ref": "SRC-30", "operator_id": "1"},
            {"subject_code": "3001", "debit": "0", "credit": "8.00", "note": "equity", "set_id": "SRC-30", "source": "generated", "rule": "UP_INV", "evidence_ref": "SRC-30", "operator_id": "1"},
        ]
        with self.engine.begin() as conn:
            conn.execute(
                text(
                    """
                    INSERT INTO consolidation_adjustments (
                        group_id, period, status, operator_id, lines_json,
                        source, tag, rule_code, evidence_ref, batch_id, note
                    ) VALUES (
                        :gid, '2025-04', 'draft', 1, :lines_json,
                        'generated', 'src', 'UP_INV', 'SRC-30', 'SRC-30', 'src-30'
                    )
                    """
                ),
                {"gid": gid, "lines_json": json.dumps(lines, ensure_ascii=False)},
            )

    def test_01_generate_disclosure_and_audit_package(self):
        gid = self._create_group()
        self._insert_source_adjustment(gid)
        payload = {"consolidation_group_id": gid, "period": "2025-04", "operator_id": 1}

        self.assertEqual(self.client.post("/task/cons-24", json=payload).status_code, 200)
        self.assertEqual(self.client.post("/task/cons-28", json=payload).status_code, 200)
        self.assertEqual(self.client.post("/task/cons-29", json={**payload, "approver_id": 99, "auto_approve": True}).status_code, 200)

        first = self.client.post("/task/cons-30", json=payload)
        self.assertEqual(first.status_code, 200, first.get_data(as_text=True))
        body = first.get_json() or {}
        self.assertEqual(str(body.get("status") or ""), "success")
        self.assertEqual(str(body.get("message") or ""), "报表披露与审计包生成完成")
        self.assertEqual(str(body.get("rule_code") or ""), "CONS30_DISCLOSURE_AUDIT_PACKAGE")
        self.assertEqual(int(body.get("report_count") or 0), 4)
        self.assertTrue(str(body.get("file_name") or "").endswith(".xlsx"))
        self.assertTrue(len(body.get("sheet_names") or []) >= 6)

        second = self.client.post("/task/cons-30", json=payload)
        self.assertEqual(second.status_code, 200, second.get_data(as_text=True))
        self.assertEqual(int((second.get_json() or {}).get("package_id") or 0), int(body.get("package_id") or 0))

        with self.engine.connect() as conn:
            row = conn.execute(
                text(
                    """
                    SELECT file_name, source, rule_code, status, package_meta_json, package_blob
                    FROM consolidation_audit_packages
                    WHERE group_id=:gid AND period='2025-04'
                    ORDER BY id DESC
                    LIMIT 1
                    """
                ),
                {"gid": gid},
            ).fetchone()
        self.assertIsNotNone(row)
        self.assertEqual(str(row.source or ""), "generated")
        self.assertEqual(str(row.rule_code or ""), "CONS30_DISCLOSURE_AUDIT_PACKAGE")
        self.assertEqual(str(row.status or ""), "draft")
        meta = json.loads(str(row.package_meta_json or "{}"))
        self.assertEqual(int(meta.get("report_count") or 0), 4)
        blob = bytes(row.package_blob or b"")
        self.assertTrue(len(blob) > 0)

        wb = load_workbook(io.BytesIO(blob), data_only=False)
        expected_sheets = {"Disclosure_Index", "Balance_Sheet", "Income_Statement", "Cash_Flow", "Equity_Change", "Audit_Trail"}
        self.assertTrue(expected_sheets.issubset(set(wb.sheetnames)))
        self.assertEqual(str(wb["Disclosure_Index"]["B5"].value or ""), "=COUNTA(Audit_Trail!A:A)-1")
        self.assertEqual(str(wb["Balance_Sheet"]["C20"].value or ""), "=SUM(C3:C19)")
        self.assertEqual(str(wb["Audit_Trail"]["A1"].value or ""), "index_id")


if __name__ == "__main__":
    unittest.main()
