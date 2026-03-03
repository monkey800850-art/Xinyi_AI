import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from app.services.export_service import export_report


class VoucherJournalExportTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        fd, path = tempfile.mkstemp(prefix="voucher_journal_export_", suffix=".db")
        os.close(fd)
        cls.db_path = Path(path)
        cls.engine = create_engine(f"sqlite:///{cls.db_path}", future=True)
        with cls.engine.begin() as conn:
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS vouchers (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        book_id INTEGER NOT NULL,
                        voucher_date DATE NOT NULL,
                        voucher_word VARCHAR(16) NULL,
                        voucher_no VARCHAR(32) NULL,
                        status VARCHAR(16) NOT NULL
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS voucher_lines (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        voucher_id INTEGER NOT NULL,
                        line_no INTEGER NOT NULL,
                        summary VARCHAR(255) NULL,
                        subject_code VARCHAR(64) NOT NULL,
                        subject_name VARCHAR(255) NOT NULL,
                        aux_display VARCHAR(255) NULL,
                        debit NUMERIC(18,2) NOT NULL DEFAULT 0,
                        credit NUMERIC(18,2) NOT NULL DEFAULT 0,
                        note VARCHAR(255) NULL
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS export_audit_logs (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        report_key VARCHAR(64) NOT NULL,
                        book_id INTEGER NULL,
                        filters TEXT NULL,
                        file_name VARCHAR(255) NOT NULL,
                        operator VARCHAR(64) NULL
                    )
                    """
                )
            )

    @classmethod
    def tearDownClass(cls):
        try:
            cls.engine.dispose()
        finally:
            if cls.db_path.exists():
                cls.db_path.unlink()

    def setUp(self):
        with self.engine.begin() as conn:
            conn.execute(text("DELETE FROM voucher_lines"))
            conn.execute(text("DELETE FROM vouchers"))
            conn.execute(text("DELETE FROM export_audit_logs"))
            conn.execute(
                text(
                    """
                    INSERT INTO vouchers (id, book_id, voucher_date, voucher_word, voucher_no, status) VALUES
                    (1, 1, '2026-03-01', '记', '0001', 'posted'),
                    (2, 1, '2026-03-01', '记', '0002', 'draft')
                    """
                )
            )
            conn.execute(
                text(
                    """
                    INSERT INTO voucher_lines (voucher_id, line_no, summary, subject_code, subject_name, aux_display, debit, credit, note) VALUES
                    (1, 1, '销售收入', '6001', '主营业务收入', '', 0.00, 100.00, '已过账'),
                    (1, 2, '销售收入', '1122', '应收账款', 'E001 客户A', 100.00, 0.00, ''),
                    (2, 1, '草稿分录', '1001', '库存现金', '', 10.00, 0.00, '草稿')
                    """
                )
            )

    def test_export_voucher_journal_xlsx_default_posted_only(self):
        params = {"book_id": "1", "start_date": "2026-03-01", "end_date": "2026-03-31"}
        with patch("app.services.export_service.get_engine", return_value=self.engine), patch(
            "app.services.export_service.log_audit"
        ):
            content, file_name = export_report("voucher_journal", params, operator="tester")

        self.assertTrue(file_name.startswith("voucher_journal_"))
        self.assertTrue(file_name.endswith(".xlsx"))
        self.assertGreater(len(content), 200)

        from io import BytesIO

        loaded = load_workbook(BytesIO(content))
        ws = loaded.active
        self.assertEqual(ws["A1"].value, "序时账（会计分录流水）")
        self.assertEqual(ws["A5"].value, "日期")
        self.assertEqual(ws["B5"].value, "凭证号")
        self.assertEqual(ws["A6"].value, "2026-03-01")
        self.assertEqual(ws["B6"].value, "记0001")
        self.assertEqual(ws["E6"].value, "销售收入")
        self.assertEqual(ws["H7"].value, "E001 客户A")
        # default status=posted, draft voucher should not appear
        self.assertEqual(ws.max_row, 7)


if __name__ == "__main__":
    unittest.main()
